#!/usr/bin/env python3
"""
Amazon settlement PDF -> accounting XLSX filler.

This script appends or updates workbook rows from Amazon settlement PDFs while
preserving existing workbook formatting and formulas as much as possible.

Default behavior:
- Writes a new output workbook; never overwrites input unless --in-place is used.
- Copies row style/formulas from the template row before writing values.
- Writes an external audit CSV next to the output workbook.
- Uses ChinaMoney / CFETS RMB central parity rates and caches fetched rates.
"""
from __future__ import annotations

import argparse
import csv
import datetime as dt
import json
import random
import re
import sys
import time
from copy import copy
from dataclasses import dataclass, asdict
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence, Tuple

try:
    import fitz  # PyMuPDF
except Exception:  # pragma: no cover
    fitz = None

try:
    import requests
except Exception:  # pragma: no cover
    requests = None

from openpyxl import load_workbook
from openpyxl.formula.translate import Translator


CHINAMONEY_HIST_URL = "https://www.chinamoney.com.cn/ags/ms/cm-u-bk-ccpr/CcprHisNew"
CHINAMONEY_REFERER = "https://www.chinamoney.com.cn/chinese/bkccpr/"

SUMMARY_HEADERS = [
    "summaries", "zusammenfassungen", "resumen", "resúmenes", "riepilogo", "riepiloghi",
    "synthèse", "syntheses", "synthèses", "samenvattingen", "podsumowanie", "概要", "サマリー",
]

LABELS = {
    "income": ["income", "einnahmen", "ingresos", "entrate", "revenus", "recettes", "inkomsten", "przychody", "receitas", "gelir", "売上", "収入", "收入"],
    "expenses": ["expenses", "ausgaben", "gastos", "spese", "dépenses", "depenses", "kosten", "uitgaven", "wydatki", "despesas", "gider", "費用", "支出"],
    "tax": ["tax", "steuer", "impuesto", "imposte", "taxes", "belasting", "podatek", "imposto", "vergi", "税", "稅"],
    "transfer": ["transfers", "transfer to bank account", "transfer", "übertragungen", "uebertragungen", "überweisungen auf bankkonto", "ueberweisungen auf bankkonto", "transferencias", "trasferimenti", "virements", "transferts", "overboekingen", "przelewy", "pagamentos", "bankkonto", "bank account", "振込", "銀行", "转账", "轉賬", "回款", "入金"],
}

CURRENCY_FROM_SITE = {
    "US": "USD", "CA": "CAD", "MX": "MXN", "BR": "BRL", "UK": "GBP", "GB": "GBP",
    "DE": "EUR", "FR": "EUR", "IT": "EUR", "ES": "EUR", "NL": "EUR", "BE": "EUR", "IE": "EUR",
    "SE": "SEK", "PL": "PLN", "TR": "TRY", "JP": "JPY", "AU": "AUD", "SG": "SGD",
    "AE": "AED", "SA": "SAR", "IN": "INR",
}

CURRENCY_TEXT_PATTERNS = [
    (r"\bEuro\b|\bEUR\b|euros", "EUR"),
    (r"US Dollars?|U\.S\. Dollars?|USD|Dólares estadounidenses|Dollar américain", "USD"),
    (r"British Pounds?|Pound Sterling|GBP|livres sterling", "GBP"),
    (r"Mexican Pesos?|MXN|peso mexicano", "MXN"),
    (r"Canadian Dollars?|CAD", "CAD"),
    (r"Australian Dollars?|AUD", "AUD"),
    (r"Japanese Yen|JPY|Yen", "JPY"),
    (r"Polish Zloty|Polish Złoty|PLN|Zloty", "PLN"),
    (r"Swedish Krona|SEK|Krona", "SEK"),
    (r"Turkish Lira|TRY|Lira", "TRY"),
    (r"Singapore Dollars?|SGD", "SGD"),
]

MONTHS = {
    "jan": 1, "january": 1, "janvier": 1, "enero": 1, "gennaio": 1, "januar": 1,
    "feb": 2, "february": 2, "février": 2, "fevrier": 2, "febrero": 2, "febbraio": 2, "februar": 2,
    "mar": 3, "march": 3, "märz": 3, "maerz": 3, "mars": 3, "marzo": 3,
    "apr": 4, "april": 4, "avril": 4, "abril": 4, "aprile": 4,
    "may": 5, "mai": 5, "mayo": 5, "maggio": 5,
    "jun": 6, "june": 6, "juni": 6, "juin": 6, "junio": 6, "giugno": 6,
    "jul": 7, "july": 7, "juli": 7, "juillet": 7, "julio": 7, "luglio": 7,
    "aug": 8, "august": 8, "août": 8, "aout": 8, "agosto": 8,
    "sep": 9, "sept": 9, "september": 9, "septembre": 9, "septiembre": 9, "settembre": 9,
    "oct": 10, "october": 10, "okt": 10, "oktober": 10, "octobre": 10, "octubre": 10, "ottobre": 10,
    "nov": 11, "november": 11, "novembre": 11, "noviembre": 11,
    "dec": 12, "december": 12, "dez": 12, "dezember": 12, "décembre": 12, "decembre": 12, "diciembre": 12, "dicembre": 12,
}

HEADER_ALIASES = {
    "month": ["时间", "月份", "日期", "month", "date"],
    "platform": ["平台", "platform"],
    "store_site": ["店铺(按站点)", "店铺按站点", "店铺", "站点", "store", "site"],
    "registered_company": ["注册主体(公司名字)", "注册主体", "公司名字", "公司名称", "company", "legal entity"],
    "report_filename": ["亚马逊结算报告", "结算报告", "settlement report", "report"],
    "income": ["店铺销售额（原币）", "店铺销售额(原币)", "销售额原币", "income", "sales original"],
    "tax": ["销售税tax（原币）", "销售税tax(原币)", "销售税", "tax"],
    "income_tax_included": ["店铺销售额（原币包含tax）", "店铺销售额(原币包含tax)", "包含tax", "含tax"],
    "expenses": ["销售费用（原币）", "销售费用(原币)", "销售费用", "expenses"],
    "transfer": ["账单回款额（原币）", "账单回款额(原币)", "账单回款额", "transfer", "bank transfer"],
    "currency": ["币种", "currency"],
    "rate": ["汇率", "exchange rate", "rate"],
}


@dataclass
class PdfExtraction:
    pdf_path: str
    filename: str
    month_start: str
    store_site: str
    display_name: str
    registered_company: str
    currency: str
    income_raw: Decimal
    expenses_raw: Decimal
    tax_raw: Optional[Decimal]
    transfer_raw: Decimal
    income_write: Decimal
    expenses_write: Decimal
    tax_write: Optional[Decimal]
    transfer_write: Decimal
    extraction_notes: str = ""


@dataclass
class RateInfo:
    currency: str
    target_date: str
    source_date: str
    raw_pair: str
    raw_value: Decimal
    workbook_rate: Decimal
    source: str


@dataclass
class AuditRow:
    pdf_file: str
    row: int
    month_start: str
    store_site: str
    registered_company: str
    currency: str
    income_pdf_raw: str
    expenses_pdf_raw: str
    tax_pdf_raw: str
    transfer_pdf_raw: str
    income_written: str
    expenses_written: str
    tax_written: str
    transfer_written: str
    rate: str
    rate_source_date: str
    rate_source: str
    status: str
    notes: str


def q2(value: Optional[Decimal]) -> str:
    if value is None:
        return ""
    return str(value.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))


def parse_amount(token: str) -> Decimal:
    s = token.strip().replace("\u00a0", " ").replace(" ", "")
    neg = False
    if s.startswith("(") and s.endswith(")"):
        neg = True
        s = s[1:-1]
    if s.startswith("−"):
        neg = True
        s = s[1:]
    if s.startswith("-"):
        neg = True
        s = s[1:]
    s = re.sub(r"[^0-9,.-]", "", s)
    if not s:
        raise ValueError(f"No numeric amount in {token!r}")
    last_comma = s.rfind(",")
    last_dot = s.rfind(".")
    if last_comma > last_dot:
        s = s.replace(".", "").replace(",", ".")
    elif last_dot > last_comma:
        s = s.replace(",", "")
    else:
        if "," in s:
            s = s.replace(",", ".")
    try:
        val = Decimal(s)
    except InvalidOperation as e:
        raise ValueError(f"Cannot parse amount {token!r}") from e
    return -val if neg else val


AMOUNT_RE = re.compile(r"(?<!\w)[-−(]?\d{1,3}(?:[.,]\d{3})*(?:[.,]\d{2,4})\)?|(?<!\w)[-−(]?\d+(?:[.,]\d{2,4})\)?")


def normalize_text(text: str) -> str:
    t = text.lower().replace("\u00a0", " ")
    t = t.replace("ü", "u").replace("ä", "a").replace("ö", "o").replace("ß", "ss")
    t = re.sub(r"\s+", " ", t)
    return t.strip()


def row_groups_from_pdf(pdf_path: Path) -> Tuple[str, List[Dict[str, Any]]]:
    if fitz is None:
        raise RuntimeError("PyMuPDF is required. Install with `pip install pymupdf`.")
    doc = fitz.open(str(pdf_path))
    all_text_parts: List[str] = []
    rows: List[Dict[str, Any]] = []
    for page_no, page in enumerate(doc):
        all_text_parts.append(page.get_text())
        words = page.get_text("words")
        buckets: List[List[Any]] = []
        for w in sorted(words, key=lambda x: (page_no, x[1], x[0])):
            placed = False
            for bucket in buckets:
                if abs(bucket[0][1] - w[1]) <= 3.0:
                    bucket.append(w)
                    placed = True
                    break
            if not placed:
                buckets.append([w])
        for bucket in buckets:
            bucket_sorted = sorted(bucket, key=lambda x: x[0])
            text = " ".join(w[4] for w in bucket_sorted)
            rows.append({"page": page_no + 1, "y": sum(w[1] for w in bucket_sorted) / len(bucket_sorted), "x0": min(w[0] for w in bucket_sorted), "x1": max(w[2] for w in bucket_sorted), "text": text, "words": bucket_sorted})
    return "\n".join(all_text_parts), rows


def contains_any(text: str, labels: Sequence[str]) -> bool:
    n = normalize_text(text)
    return any(normalize_text(label) in n for label in labels)


def rightmost_amount(row: Dict[str, Any]) -> Optional[Decimal]:
    candidates: List[Tuple[float, str]] = []
    for w in row["words"]:
        token = w[4]
        if AMOUNT_RE.fullmatch(token.strip()):
            try:
                parse_amount(token)
                candidates.append((w[0], token))
            except ValueError:
                pass
    if not candidates:
        matches = AMOUNT_RE.findall(row["text"])
        if matches:
            return parse_amount(matches[-1])
        return None
    candidates.sort(key=lambda x: x[0])
    return parse_amount(candidates[-1][1])


def find_summary_rows(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    summary_candidates = [r for r in rows if r["page"] == 1 and contains_any(r["text"], SUMMARY_HEADERS)]
    if summary_candidates:
        y = min(r["y"] for r in summary_candidates)
        return [r for r in rows if r["page"] == 1 and y <= r["y"] <= y + 95]
    page1 = [r for r in rows if r["page"] == 1]
    if not page1:
        return rows
    max_y = max(r["y"] for r in page1)
    return [r for r in page1 if r["y"] <= max_y * 0.35]


def extract_amount_from_rows(rows: List[Dict[str, Any]], key: str) -> Tuple[Decimal, str]:
    labels = LABELS[key]
    matches: List[Tuple[Decimal, str, float]] = []
    for row in rows:
        if contains_any(row["text"], labels):
            amount = rightmost_amount(row)
            if amount is not None:
                matches.append((amount, row["text"], row["y"]))
    if not matches:
        raise ValueError(f"Could not find {key} amount in summary rows.")
    matches.sort(key=lambda x: x[2])
    return matches[0][0], matches[0][1]


def extract_detailed_bank_transfer(rows: List[Dict[str, Any]]) -> Optional[Decimal]:
    labels = ["transfer to bank account", "überweisungen auf bankkonto", "ueberweisungen auf bankkonto", "transferencias a cuenta bancaria", "trasferimenti su conto bancario", "virements sur compte bancaire", "bank account", "bankkonto"]
    candidates = []
    for row in rows:
        if contains_any(row["text"], labels):
            amount = rightmost_amount(row)
            if amount is not None:
                candidates.append((amount, row["text"], row["page"], row["y"]))
    if not candidates:
        return None
    candidates.sort(key=lambda x: abs(x[0]), reverse=True)
    return candidates[0][0]


def parse_month_from_filename(filename: str) -> Optional[dt.date]:
    m = re.search(r"(20\d{2})[-_\.](0[1-9]|1[0-2])", filename)
    if m:
        return dt.date(int(m.group(1)), int(m.group(2)), 1)
    return None


def parse_month_from_text(text: str) -> Optional[dt.date]:
    m = re.search(r"\b([A-Za-zÀ-ÿ]{3,10})\s+\d{1,2},?\s+(20\d{2})\b", text)
    if m:
        month_name = normalize_text(m.group(1))
        month = MONTHS.get(month_name[:3], MONTHS.get(month_name))
        if month:
            return dt.date(int(m.group(2)), month, 1)
    m = re.search(r"\b(20\d{2})[-/\.](0?[1-9]|1[0-2])[-/\.]\d{1,2}\b", text)
    if m:
        return dt.date(int(m.group(1)), int(m.group(2)), 1)
    return None


def parse_store_site_from_filename(filename: str) -> str:
    stem = Path(filename).stem
    base = re.sub(r"_?Standard[-_].*$", "", stem, flags=re.I)
    base = re.sub(r"[-_]?Standard$", "", base, flags=re.I)
    if base.startswith("CR-") and len(base.split("-")) >= 3:
        base = base[3:]
    return base.strip()


def site_code_from_store_site(store_site: str) -> Optional[str]:
    parts = re.split(r"[-_ ]+", store_site.strip())
    for part in reversed(parts):
        p = part.upper()
        if p in CURRENCY_FROM_SITE:
            return p
    return None


def extract_currency(text: str, store_site: str) -> str:
    for pattern, code in CURRENCY_TEXT_PATTERNS:
        if re.search(pattern, text, flags=re.I):
            return code
    site = site_code_from_store_site(store_site)
    if site and site in CURRENCY_FROM_SITE:
        return CURRENCY_FROM_SITE[site]
    raise ValueError("Could not infer currency from PDF text or filename/site code.")


def extract_value_after_label_from_rows(rows: List[Dict[str, Any]], labels: Sequence[str]) -> str:
    for row in rows:
        if contains_any(row["text"], labels):
            text = row["text"]
            if ":" in text:
                tail = text.split(":", 1)[1].strip()
                if tail:
                    return tail
            n = normalize_text(text)
            for label in labels:
                ln = normalize_text(label)
                idx = n.find(ln)
                if idx >= 0:
                    return text[idx + len(label):].strip()
    return ""


def load_mapping(mapping_path: Optional[Path]) -> Dict[str, Dict[str, str]]:
    mapping: Dict[str, Dict[str, str]] = {}
    if not mapping_path or not mapping_path.exists():
        return mapping
    with mapping_path.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        for row in reader:
            key = (row.get("store_site") or "").strip()
            if key:
                mapping[key] = {k: (v or "").strip() for k, v in row.items()}
    return mapping


def extract_pdf(pdf_path: Path, mapping: Dict[str, Dict[str, str]]) -> PdfExtraction:
    text, rows = row_groups_from_pdf(pdf_path)
    summary_rows = find_summary_rows(rows)
    month = parse_month_from_filename(pdf_path.name) or parse_month_from_text(text)
    if not month:
        raise ValueError(f"Could not determine report month for {pdf_path.name}")
    store_site = parse_store_site_from_filename(pdf_path.name)
    map_row = mapping.get(store_site, {})
    display_name = map_row.get("display_name") or extract_value_after_label_from_rows(rows, ["Display name", "Anzeigename", "Nombre para mostrar", "Nome visualizzato", "Nom d'affichage"])
    registered_company = map_row.get("registered_company") or extract_value_after_label_from_rows(rows, ["Registered business name", "Legal entity name", "Eingetragener Firmenname", "Ragione sociale", "Razón social", "Nom légal"])
    currency = map_row.get("currency") or extract_currency(text, store_site)
    notes = []
    income_raw, _ = extract_amount_from_rows(summary_rows, "income")
    expenses_raw, _ = extract_amount_from_rows(summary_rows, "expenses")
    transfer_raw, _ = extract_amount_from_rows(summary_rows, "transfer")
    try:
        tax_raw, _ = extract_amount_from_rows(summary_rows, "tax")
    except ValueError:
        tax_raw = None
        notes.append("tax not found; left blank/0 according to workbook rule")
    detailed_transfer = extract_detailed_bank_transfer(rows)
    if detailed_transfer is not None and abs(detailed_transfer - transfer_raw) > Decimal("0.01"):
        raise ValueError(f"Transfer mismatch in {pdf_path.name}: summary={transfer_raw}, detailed bank transfer={detailed_transfer}")
    return PdfExtraction(str(pdf_path), pdf_path.name, month.isoformat(), store_site, display_name, registered_company, currency, income_raw, expenses_raw, tax_raw, transfer_raw, income_raw, abs(expenses_raw), tax_raw, abs(transfer_raw), "; ".join(notes))


def read_rate_cache(cache_path: Path) -> Dict[Tuple[str, str], RateInfo]:
    cache: Dict[Tuple[str, str], RateInfo] = {}
    if not cache_path.exists():
        return cache
    with cache_path.open("r", encoding="utf-8-sig", newline="") as f:
        for row in csv.DictReader(f):
            try:
                info = RateInfo(row["currency"], row["target_date"], row["source_date"], row.get("raw_pair", ""), Decimal(row["raw_value"]), Decimal(row["workbook_rate"]), row.get("source", "cache"))
                cache[(info.currency, info.target_date)] = info
            except Exception:
                continue
    return cache


def write_rate_cache(cache_path: Path, cache: Dict[Tuple[str, str], RateInfo]) -> None:
    cache_path.parent.mkdir(parents=True, exist_ok=True)
    with cache_path.open("w", encoding="utf-8-sig", newline="") as f:
        fields = ["currency", "target_date", "source_date", "raw_pair", "raw_value", "workbook_rate", "source"]
        writer = csv.DictWriter(f, fieldnames=fields)
        writer.writeheader()
        for _, info in sorted(cache.items()):
            writer.writerow({"currency": info.currency, "target_date": info.target_date, "source_date": info.source_date, "raw_pair": info.raw_pair, "raw_value": str(info.raw_value), "workbook_rate": str(info.workbook_rate), "source": info.source})


def parse_manual_rates(items: Sequence[str]) -> Dict[Tuple[str, str], Decimal]:
    out: Dict[Tuple[str, str], Decimal] = {}
    for item in items:
        if "=" not in item:
            raise ValueError(f"Bad --manual-rate {item!r}; expected EUR=7.9 or EUR:2026-04-01=7.9")
        left, right = item.split("=", 1)
        if ":" in left:
            cur, date = left.split(":", 1)
        else:
            cur, date = left, "*"
        out[(cur.upper().strip(), date.strip())] = Decimal(right.strip())
    return out


def fetch_chinamoney_records(start_date: dt.date, end_date: dt.date, currency: str) -> List[Dict[str, Any]]:
    if requests is None:
        raise RuntimeError("requests is required. Install with `pip install requests`.")
    params = {"startDate": start_date.strftime("%Y-%m-%d"), "endDate": end_date.strftime("%Y-%m-%d"), "currency": currency, "pageNum": "1", "pageSize": "300"}
    headers = {"Referer": CHINAMONEY_REFERER, "Origin": "https://www.chinamoney.com.cn", "X-Requested-With": "XMLHttpRequest", "User-Agent": random.choice(["Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/121 Safari/537.36", "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 Safari/605.1.15"])}
    last_error = None
    for attempt in range(3):
        try:
            resp = requests.post(CHINAMONEY_HIST_URL, params=params, headers=headers, timeout=20)
            if resp.status_code >= 400:
                resp = requests.get(CHINAMONEY_HIST_URL, params=params, headers=headers, timeout=20)
            resp.raise_for_status()
            data = resp.json()
            records = data.get("records") or data.get("data") or []
            if isinstance(records, dict):
                records = records.get("records", [])
            return list(records)
        except Exception as e:
            last_error = e
            time.sleep(1 + attempt)
    raise RuntimeError(f"ChinaMoney request failed for {currency}: {last_error}")


def rate_pair_candidates(currency: str) -> List[str]:
    cur = currency.upper()
    if cur == "JPY":
        return ["JPY/CNY", "100JPY/CNY"]
    return [f"{cur}/CNY"]


def extract_rate_value(record: Dict[str, Any]) -> Decimal:
    if "values" in record and record["values"]:
        return Decimal(str(record["values"][0]).replace(",", ""))
    for key in ["value", "rate", "price", "middlePrice", "centralParity"]:
        if key in record and record[key] not in (None, ""):
            return Decimal(str(record[key]).replace(",", ""))
    raise ValueError(f"Could not extract rate value from record: {record}")


def extract_rate_date(record: Dict[str, Any]) -> dt.date:
    for key in ["date", "showDate", "publishDate", "tradeDate"]:
        if key in record and record[key]:
            return dt.datetime.strptime(str(record[key])[:10], "%Y-%m-%d").date()
    raise ValueError(f"Could not extract date from record: {record}")


def get_rate(currency: str, target_date: dt.date, cache_path: Path, manual_rates: Dict[Tuple[str, str], Decimal], offline: bool = False) -> RateInfo:
    cur = currency.upper()
    if cur in {"CNY", "RMB"}:
        return RateInfo(cur, target_date.isoformat(), target_date.isoformat(), "CNY/CNY", Decimal("1"), Decimal("1"), "constant")
    manual = manual_rates.get((cur, target_date.isoformat())) or manual_rates.get((cur, "*"))
    if manual is not None:
        return RateInfo(cur, target_date.isoformat(), target_date.isoformat(), f"{cur}/CNY", manual, manual, "manual")
    cache = read_rate_cache(cache_path)
    cached = cache.get((cur, target_date.isoformat()))
    if cached:
        return cached
    if offline:
        raise RuntimeError(f"No cached/manual rate for {cur} on {target_date}; offline mode enabled.")
    start_date = target_date - dt.timedelta(days=12)
    best: Optional[Tuple[dt.date, Decimal, str]] = None
    errors = []
    for pair in rate_pair_candidates(cur):
        try:
            records = fetch_chinamoney_records(start_date, target_date, pair)
            for rec in records:
                d = extract_rate_date(rec)
                if d <= target_date:
                    v = extract_rate_value(rec)
                    if best is None or d > best[0]:
                        best = (d, v, pair)
        except Exception as e:
            errors.append(f"{pair}: {e}")
    if best is None:
        raise RuntimeError(f"No ChinaMoney rate found for {cur} <= {target_date}. Errors: {'; '.join(errors)}")
    source_date, raw_value, pair = best
    workbook_rate = raw_value / Decimal("100") if cur == "JPY" else raw_value
    info = RateInfo(cur, target_date.isoformat(), source_date.isoformat(), pair, raw_value, workbook_rate, "ChinaMoney/CFETS")
    cache[(cur, target_date.isoformat())] = info
    write_rate_cache(cache_path, cache)
    return info


def normalize_header(s: Any) -> str:
    if s is None:
        return ""
    t = str(s).lower().replace("\n", "")
    t = re.sub(r"[\s　]+", "", t).replace("（", "(").replace("）", ")")
    return t


def find_header_row_and_cols(ws) -> Tuple[int, Dict[str, int]]:
    normalized_aliases = {key: [normalize_header(a) for a in aliases] for key, aliases in HEADER_ALIASES.items()}
    best_row = None
    best_score = -1
    best_cols: Dict[str, int] = {}
    for row in range(1, min(ws.max_row, 10) + 1):
        values = [ws.cell(row, col).value for col in range(1, ws.max_column + 1)]
        cols: Dict[str, int] = {}
        for col, val in enumerate(values, start=1):
            hv = normalize_header(val)
            if not hv:
                continue
            for key, aliases in normalized_aliases.items():
                if key in cols:
                    continue
                if any(alias and (alias == hv or alias in hv or hv in alias) for alias in aliases):
                    cols[key] = col
        if len(cols) > best_score:
            best_row, best_score, best_cols = row, len(cols), cols
    required = ["month", "platform", "store_site", "registered_company", "report_filename", "income", "expenses", "transfer", "currency", "rate"]
    missing = [key for key in required if key not in best_cols]
    if missing:
        raise ValueError(f"Could not match required workbook headers: {missing}. Matched columns: {best_cols}")
    return int(best_row), best_cols


def copy_row(ws, src_row: int, dst_row: int) -> None:
    if src_row == dst_row:
        return
    ws.row_dimensions[dst_row].height = ws.row_dimensions[src_row].height
    for col in range(1, ws.max_column + 1):
        src = ws.cell(src_row, col)
        dst = ws.cell(dst_row, col)
        if src.has_style:
            dst._style = copy(src._style)
        if src.number_format:
            dst.number_format = src.number_format
        if src.font:
            dst.font = copy(src.font)
        if src.fill:
            dst.fill = copy(src.fill)
        if src.border:
            dst.border = copy(src.border)
        if src.alignment:
            dst.alignment = copy(src.alignment)
        if src.protection:
            dst.protection = copy(src.protection)
        if src.data_type == "f" or (isinstance(src.value, str) and src.value.startswith("=")):
            try:
                dst.value = Translator(str(src.value), origin=src.coordinate).translate_formula(dst.coordinate)
            except Exception:
                dst.value = src.value
        else:
            dst.value = None


def find_existing_or_next_row(ws, header_row: int, cols: Dict[str, int], filename: str, template_row: int) -> int:
    report_col = cols["report_filename"]
    for row in range(header_row + 1, ws.max_row + 1):
        value = ws.cell(row, report_col).value
        if value and str(value).strip() == filename:
            return row
    for row in range(header_row + 1, ws.max_row + 2):
        if not ws.cell(row, report_col).value:
            if row > ws.max_row:
                copy_row(ws, template_row, row)
            return row
    return ws.max_row + 1


def to_excel_date(date_iso: str) -> dt.datetime:
    d = dt.date.fromisoformat(date_iso)
    return dt.datetime(d.year, d.month, d.day)


def write_decimal_cell(cell, value: Optional[Decimal]) -> None:
    cell.value = None if value is None else float(value)


def set_calc_mode(wb) -> None:
    try:
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
        wb.calculation.calcMode = "auto"
    except Exception:
        pass


def maybe_recalculate_with_excel(path: Path) -> bool:
    try:
        import xlwings as xw  # type: ignore
    except Exception:
        return False
    app = None
    try:
        app = xw.App(visible=False, add_book=False)
        book = app.books.open(str(path.resolve()))
        app.calculate()
        book.save()
        book.close()
        return True
    except Exception:
        return False
    finally:
        if app is not None:
            try:
                app.quit()
            except Exception:
                pass


def write_workbook(workbook_path: Path, output_path: Path, extractions: List[PdfExtraction], rates: Dict[str, RateInfo], dry_run: bool, recalc_excel: bool) -> List[AuditRow]:
    wb = load_workbook(workbook_path)
    ws = wb.active
    header_row, cols = find_header_row_and_cols(ws)
    template_row = header_row + 1
    audit: List[AuditRow] = []
    for ext in extractions:
        row = find_existing_or_next_row(ws, header_row, cols, ext.filename, template_row)
        if row > ws.max_row or all(ws.cell(row, c).value is None for c in range(1, ws.max_column + 1)):
            copy_row(ws, template_row, row)
        rate = rates[ext.filename]
        if not dry_run:
            ws.cell(row, cols["month"]).value = to_excel_date(ext.month_start)
            ws.cell(row, cols["platform"]).value = "亚马逊"
            ws.cell(row, cols["store_site"]).value = ext.store_site
            ws.cell(row, cols["registered_company"]).value = ext.registered_company
            ws.cell(row, cols["report_filename"]).value = ext.filename
            write_decimal_cell(ws.cell(row, cols["income"]), ext.income_write)
            if "tax" in cols:
                write_decimal_cell(ws.cell(row, cols["tax"]), ext.tax_write if ext.tax_write is not None else Decimal("0"))
            write_decimal_cell(ws.cell(row, cols["expenses"]), ext.expenses_write)
            write_decimal_cell(ws.cell(row, cols["transfer"]), ext.transfer_write)
            ws.cell(row, cols["currency"]).value = ext.currency
            write_decimal_cell(ws.cell(row, cols["rate"]), rate.workbook_rate)
        status = "OK"
        notes = ext.extraction_notes
        def cell_decimal(col_key: str) -> Optional[Decimal]:
            v = ws.cell(row, cols[col_key]).value
            if v is None:
                return None
            if isinstance(v, (int, float)):
                return Decimal(str(v))
            try:
                return Decimal(str(v))
            except Exception:
                return None
        if not dry_run:
            checks = [("income", ext.income_write), ("expenses", ext.expenses_write), ("transfer", ext.transfer_write), ("rate", rate.workbook_rate)]
            for key, expected in checks:
                actual = cell_decimal(key)
                if actual is None or abs(actual - expected) > Decimal("0.01"):
                    status = "VERIFY_FAILED"
                    notes += f"; {key} cell mismatch expected {expected} actual {actual}"
        audit.append(AuditRow(ext.filename, row, ext.month_start, ext.store_site, ext.registered_company, ext.currency, q2(ext.income_raw), q2(ext.expenses_raw), q2(ext.tax_raw), q2(ext.transfer_raw), q2(ext.income_write), q2(ext.expenses_write), q2(ext.tax_write), q2(ext.transfer_write), str(rate.workbook_rate), rate.source_date, rate.source, status, notes))
    if not dry_run:
        set_calc_mode(wb)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)
        if recalc_excel:
            maybe_recalculate_with_excel(output_path)
    return audit


def write_audit_csv(path: Path, audit: List[AuditRow]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    fields = list(asdict(audit[0]).keys()) if audit else ["pdf_file", "row", "month_start", "store_site", "registered_company", "currency", "income_pdf_raw", "expenses_pdf_raw", "tax_pdf_raw", "transfer_pdf_raw", "income_written", "expenses_written", "tax_written", "transfer_written", "rate", "rate_source_date", "rate_source", "status", "notes"]
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fields)
        writer.writeheader()
        for item in audit:
            writer.writerow(asdict(item))


def collect_pdfs(pdf_dir: Optional[Path], pdf_files: Sequence[Path]) -> List[Path]:
    found: List[Path] = []
    if pdf_dir:
        found.extend(sorted(pdf_dir.glob("*.pdf")))
    found.extend(pdf_files)
    seen = set()
    out = []
    for p in found:
        rp = p.resolve()
        if rp not in seen:
            seen.add(rp)
            out.append(p)
    return out


def main(argv: Optional[Sequence[str]] = None) -> int:
    parser = argparse.ArgumentParser(description="Fill accounting XLSX from Amazon settlement PDFs.")
    parser.add_argument("--workbook", required=True, type=Path, help="Input .xlsx workbook/template.")
    parser.add_argument("--pdf-dir", type=Path, help="Directory containing PDF files.")
    parser.add_argument("--pdf", dest="pdf_files", action="append", default=[], type=Path, help="Individual PDF file; can be repeated.")
    parser.add_argument("--output", type=Path, help="Output .xlsx path. Defaults to <input>_filled.xlsx")
    parser.add_argument("--in-place", action="store_true", help="Overwrite the input workbook. Not recommended.")
    parser.add_argument("--mapping", type=Path, help="Optional store mapping CSV.")
    parser.add_argument("--rate-cache", type=Path, default=Path("rate_cache.csv"), help="CSV cache for ChinaMoney rates.")
    parser.add_argument("--manual-rate", action="append", default=[], help="Manual fallback rate, e.g. EUR=7.916 or EUR:2026-04-01=7.916")
    parser.add_argument("--offline", action="store_true", help="Do not call ChinaMoney; use cache/manual rates only.")
    parser.add_argument("--dry-run", action="store_true", help="Extract and audit only; do not write workbook.")
    parser.add_argument("--audit", type=Path, help="Audit CSV path. Defaults next to output workbook.")
    parser.add_argument("--recalculate-excel", action="store_true", help="Try to recalculate formulas with local Microsoft Excel via xlwings.")
    args = parser.parse_args(argv)
    if not args.workbook.exists():
        raise FileNotFoundError(args.workbook)
    pdfs = collect_pdfs(args.pdf_dir, args.pdf_files)
    if not pdfs:
        raise FileNotFoundError("No PDF files found. Use --pdf-dir or --pdf.")
    for p in pdfs:
        if not p.exists():
            raise FileNotFoundError(p)
    output = args.output
    if args.in_place:
        output = args.workbook
    elif output is None:
        output = args.workbook.with_name(args.workbook.stem + "_filled" + args.workbook.suffix)
    audit_path = args.audit or output.with_suffix(".audit.csv")
    mapping = load_mapping(args.mapping)
    manual_rates = parse_manual_rates(args.manual_rate)
    extractions: List[PdfExtraction] = []
    rates: Dict[str, RateInfo] = {}
    for pdf in pdfs:
        ext = extract_pdf(pdf, mapping)
        extractions.append(ext)
        target_date = dt.date.fromisoformat(ext.month_start)
        rates[ext.filename] = get_rate(ext.currency, target_date, args.rate_cache, manual_rates, args.offline)
    audit = write_workbook(args.workbook, output, extractions, rates, args.dry_run, args.recalculate_excel)
    write_audit_csv(audit_path, audit)
    print(json.dumps({"status": "ok" if all(a.status == "OK" for a in audit) else "check_audit", "dry_run": args.dry_run, "output": None if args.dry_run else str(output), "audit": str(audit_path), "rows": [asdict(a) for a in audit]}, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except Exception as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        raise
