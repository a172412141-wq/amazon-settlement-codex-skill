#!/usr/bin/env python3
"""
One-step PDF -> data table workflow.

Usage examples:
  python .agents/skills/amazon-settlement-xlsx/scripts/pdf_to_table.py ./pdfs
  python .agents/skills/amazon-settlement-xlsx/scripts/pdf_to_table.py ./file1.pdf ./file2.pdf

The script creates a standalone XLSX table. It does not require an existing Excel
workbook template.
"""
from __future__ import annotations

import argparse
import csv
import datetime as dt
import json
import sys
from dataclasses import asdict
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Sequence, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from run import (
    PdfExtraction,
    RateInfo,
    collect_pdfs,
    extract_pdf,
    get_rate,
    load_mapping,
    parse_manual_rates,
    q2,
)

HEADERS = [
    "月份",
    "平台",
    "店铺(按站点)",
    "注册主体(公司名字)",
    "亚马逊结算报告",
    "币种",
    "汇率日期",
    "汇率",
    "店铺销售额（原币）",
    "销售税Tax（原币）",
    "店铺销售额（原币 包含TAX）",
    "销售费用（原币）",
    "账单回款额（原币）",
    "店铺销售额（人民币）",
    "销售税Tax（人民币）",
    "店铺销售额（人民币 包含TAX）",
    "销售费用（人民币）",
    "账单回款额（人民币）",
    "PDF原始expenses",
    "PDF原始transfer",
    "状态",
    "备注",
]


def decimal_or_none(value: Optional[Decimal]) -> Optional[Decimal]:
    return value if value is not None else None


def round_money(value: Optional[Decimal]) -> Optional[Decimal]:
    if value is None:
        return None
    return value.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def as_float(value: Optional[Decimal]) -> Optional[float]:
    if value is None:
        return None
    return float(value)


def calc_cny(value: Optional[Decimal], rate: Optional[Decimal]) -> Optional[Decimal]:
    if value is None or rate is None:
        return None
    return round_money(value * rate)


def resolve_inputs(inputs: Sequence[Path]) -> List[Path]:
    if not inputs:
        default_dir = Path("pdfs")
        if default_dir.exists():
            return collect_pdfs(default_dir, [])
        raise FileNotFoundError("No PDF input provided and ./pdfs does not exist.")
    pdfs: List[Path] = []
    for item in inputs:
        if item.is_dir():
            pdfs.extend(collect_pdfs(item, []))
        else:
            pdfs.append(item)
    unique: List[Path] = []
    seen = set()
    for pdf in pdfs:
        rp = pdf.resolve()
        if rp not in seen:
            seen.add(rp)
            unique.append(pdf)
    return unique


def default_mapping_path(explicit: Optional[Path]) -> Optional[Path]:
    if explicit:
        return explicit
    local = Path("store_mapping.local.csv")
    if local.exists():
        return local
    bundled = Path(__file__).resolve().parents[1] / "assets" / "store_mapping.csv"
    return bundled if bundled.exists() else None


def extract_all(pdfs: Sequence[Path], mapping_path: Optional[Path], rate_cache: Path, manual_rates: Dict[Tuple[str, str], Decimal], offline: bool, strict: bool) -> List[dict]:
    mapping = load_mapping(mapping_path)
    rows: List[dict] = []
    for pdf in pdfs:
        status = "OK"
        notes = ""
        rate_info: Optional[RateInfo] = None
        ext: Optional[PdfExtraction] = None
        try:
            ext = extract_pdf(pdf, mapping)
            month_date = dt.date.fromisoformat(ext.month_start)
            try:
                rate_info = get_rate(ext.currency, month_date, rate_cache, manual_rates, offline)
            except Exception as rate_exc:
                status = "RATE_MISSING"
                notes = f"Rate lookup failed: {rate_exc}"
                if strict:
                    raise
        except Exception as exc:
            if strict:
                raise
            rows.append({
                "月份": "",
                "平台": "亚马逊",
                "店铺(按站点)": "",
                "注册主体(公司名字)": "",
                "亚马逊结算报告": pdf.name,
                "币种": "",
                "汇率日期": "",
                "汇率": None,
                "店铺销售额（原币）": None,
                "销售税Tax（原币）": None,
                "店铺销售额（原币 包含TAX）": None,
                "销售费用（原币）": None,
                "账单回款额（原币）": None,
                "店铺销售额（人民币）": None,
                "销售税Tax（人民币）": None,
                "店铺销售额（人民币 包含TAX）": None,
                "销售费用（人民币）": None,
                "账单回款额（人民币）": None,
                "PDF原始expenses": None,
                "PDF原始transfer": None,
                "状态": "EXTRACT_FAILED",
                "备注": str(exc),
            })
            continue
        assert ext is not None
        rate = rate_info.workbook_rate if rate_info else None
        tax = ext.tax_write if ext.tax_write is not None else Decimal("0")
        income_with_tax = ext.income_write + tax
        rows.append({
            "月份": ext.month_start,
            "平台": "亚马逊",
            "店铺(按站点)": ext.store_site,
            "注册主体(公司名字)": ext.registered_company,
            "亚马逊结算报告": ext.filename,
            "币种": ext.currency,
            "汇率日期": rate_info.source_date if rate_info else "",
            "汇率": rate,
            "店铺销售额（原币）": ext.income_write,
            "销售税Tax（原币）": tax,
            "店铺销售额（原币 包含TAX）": income_with_tax,
            "销售费用（原币）": ext.expenses_write,
            "账单回款额（原币）": ext.transfer_write,
            "店铺销售额（人民币）": calc_cny(ext.income_write, rate),
            "销售税Tax（人民币）": calc_cny(tax, rate),
            "店铺销售额（人民币 包含TAX）": calc_cny(income_with_tax, rate),
            "销售费用（人民币）": calc_cny(ext.expenses_write, rate),
            "账单回款额（人民币）": calc_cny(ext.transfer_write, rate),
            "PDF原始expenses": ext.expenses_raw,
            "PDF原始transfer": ext.transfer_raw,
            "状态": status,
            "备注": notes or ext.extraction_notes,
        })
    return rows


def write_xlsx(rows: Sequence[dict], output: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Amazon Settlement"
    ws.append(HEADERS)
    for row_data in rows:
        ws.append([row_data.get(h) for h in HEADERS])
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.fill = header_fill
    number_cols = [
        "汇率", "店铺销售额（原币）", "销售税Tax（原币）", "店铺销售额（原币 包含TAX）",
        "销售费用（原币）", "账单回款额（原币）", "店铺销售额（人民币）", "销售税Tax（人民币）",
        "店铺销售额（人民币 包含TAX）", "销售费用（人民币）", "账单回款额（人民币）",
        "PDF原始expenses", "PDF原始transfer",
    ]
    number_indexes = {HEADERS.index(h) + 1 for h in number_cols}
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if cell.column in number_indexes:
                cell.number_format = "#,##0.00"
            cell.alignment = Alignment(vertical="center")
    widths = {
        "A": 12, "B": 10, "C": 18, "D": 26, "E": 32, "F": 10, "G": 12, "H": 12,
        "I": 16, "J": 16, "K": 22, "L": 16, "M": 16, "N": 18, "O": 18, "P": 24,
        "Q": 18, "R": 18, "S": 18, "T": 18, "U": 14, "V": 40,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)


def write_csv(rows: Sequence[dict], output: Path) -> None:
    output.parent.mkdir(parents=True, exist_ok=True)
    with output.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=HEADERS)
        writer.writeheader()
        for row in rows:
            out = {}
            for h in HEADERS:
                v = row.get(h)
                out[h] = str(v) if isinstance(v, Decimal) else v
            writer.writerow(out)


def main(argv: Optional[Sequence[str]] = None) -> int:
    parser = argparse.ArgumentParser(description="Generate a standalone settlement data table from Amazon PDF files.")
    parser.add_argument("inputs", nargs="*", type=Path, help="PDF file(s) or folder(s). If omitted, uses ./pdfs.")
    parser.add_argument("--output", "-o", type=Path, default=Path("amazon_settlement_table.xlsx"), help="Output XLSX file.")
    parser.add_argument("--csv", dest="csv_output", type=Path, help="Optional CSV output path.")
    parser.add_argument("--mapping", type=Path, help="Optional local store mapping CSV. Defaults to ./store_mapping.local.csv if present.")
    parser.add_argument("--rate-cache", type=Path, default=Path(".cache/chinamoney_rates.csv"), help="Local exchange-rate cache, ignored by Git.")
    parser.add_argument("--manual-rate", action="append", default=[], help="Manual rate, e.g. EUR=7.916 or EUR:2026-04-01=7.916")
    parser.add_argument("--offline", action="store_true", help="Use cache/manual rates only; do not call ChinaMoney.")
    parser.add_argument("--strict", action="store_true", help="Stop on the first PDF or rate error. Default writes failed rows with status.")
    args = parser.parse_args(argv)
    pdfs = resolve_inputs(args.inputs)
    if not pdfs:
        raise FileNotFoundError("No PDF files found.")
    for pdf in pdfs:
        if not pdf.exists():
            raise FileNotFoundError(pdf)
    manual_rates = parse_manual_rates(args.manual_rate)
    rows = extract_all(pdfs, default_mapping_path(args.mapping), args.rate_cache, manual_rates, args.offline, args.strict)
    write_xlsx(rows, args.output)
    if args.csv_output:
        write_csv(rows, args.csv_output)
    summary = {
        "status": "ok" if all(r.get("状态") == "OK" for r in rows) else "check_rows",
        "pdf_count": len(pdfs),
        "output": str(args.output),
        "csv": str(args.csv_output) if args.csv_output else None,
        "rows": [{k: (str(v) if isinstance(v, Decimal) else v) for k, v in row.items()} for row in rows],
    }
    print(json.dumps(summary, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except Exception as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        raise
